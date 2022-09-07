import os
import pathlib
import copy
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import torch
from sklearn.metrics import accuracy_score
import cc3d
import openpyxl
import skimage.transform as skTrans
class ModelTrainer:
    """
    A class for fitting a model.

    Parameters
    ----------
    model : a subclass of `torch.nn.Module`
        A model to fit.
    dataloaders : dict of `torch.utils.data.DataLoader`
        A dictionary with 'train' and 'val' keys specifying dataloaders for model training and validation.
    criterion : a subclass of `torch.nn.Module`
        A loss function used for model training.
    optimizer : a subclass of `torch.optim.Optimizer`
        An optimizer for training a model.
    metric : function or a subclass of `torch.nn.Module`
        A metric used for evaluation.
    mode : str
        Must be 'min' or 'max'. If 'max', a model with the highest metric will be treated as the best one.
    scheduler : a class from `torch.optim.lr_scheduler` (a subclass of _LRScheduler)
        A method to adjust a learning rate during training.
    num_epochs : int
        A number of epochs for training.
    parallel : bool
        Train a model on multiple GPUs using `torch.nn.DataParallel`.
    cuda_device : str
        A CUDA device used for training and validation. If the device is unavailable,
        all computation are performed on a CPU.
    save_last_model : bool
        If 'true', a checkpoint of the last epoch will be saved (for inference and/or resuming training).
    scheduler_step_per_epoch : bool
        If 'true', a learning rate adjustment is performed after each epoch. Otherwise, after each training batch.

    Attributes
    ----------
    learning_curves : dict of dict
        A dictionary containing train & validation learning curves (loss and metric).
     best_val_epoch : int
        Indicates an epoch with the best metric on a validation set.
    best_model_wts
        Weights of the best model.
    checkpoint
        The model weights and optimizer state after the last epoch.
    """

    def __init__(self, model, dataloaders, criterion, optimizer,
                 metric=None,recall=None,precision=None, mode='max', scheduler=None, num_epochs=25,
                 parallel=False, cuda_device="cuda:0",
                 save_last_model=True, scheduler_step_per_epoch=True,filepath=None):

        self.model = model
        self.dataloaders = dataloaders
        self.criterion = criterion
        self.metric = metric
        self.recall = recall
        self.precision = precision
        self.mode = mode
        self.optimizer = optimizer
        self.scheduler = scheduler
        self.num_epochs = num_epochs
        self.parallel = parallel
        self.device = torch.device(cuda_device if torch.cuda.is_available() else "cpu")
        self.save_last_model = save_last_model
        self.scheduler_step_per_epoch = scheduler_step_per_epoch
        self.filepath =filepath

        # Dicts for saving train and val losses:
        self.learning_curves = dict()
        self.learning_curves['loss'], self.learning_curves['metric'], self.learning_curves['recall'], \
        self.learning_curves['precision'] = dict(), dict(), dict(), dict()
        self.learning_curves['loss']['train'], self.learning_curves['loss']['val'] = [], []
        self.learning_curves['metric']['train'], self.learning_curves['metric']['val'] = [], []
        self.learning_curves['recall']['train'], self.learning_curves['recall']['val'] = [], []
        self.learning_curves['precision']['train'], self.learning_curves['precision']['val'] = [], []

        # Summary: Best epoch, loss, metric and best model weights:
        self.best_val_epoch = 0
        self.best_val_loss = float('inf')
        if self.mode == 'max':
            self.best_val_avg_metric = -float('inf')
        else:
            self.best_val_avg_metric = float('inf')
        self.best_val_metric = 0.0
        self.best_val_recall = 0.0
        self.best_val_precision = 0.0
        self.best_model_wts = None
        self.checkpoint = None  # last model and optimizer weights

    def train_model(self,path_to_dir):
        """Fit a model."""

        if self.device.type == 'cpu':
            print('Start training the model on CPU')
        elif self.parallel and torch.cuda.device_count() > 1:
            print(f'Start training the model on {torch.cuda.device_count()} '
                  f'{torch.cuda.get_device_name(torch.cuda.current_device())} in parallel')
            self.model = torch.nn.DataParallel(self.model)
        else:
            print(f'Start training the model on {torch.cuda.get_device_name(torch.cuda.current_device())}')

        self.model = self.model.to(self.device)

        for epoch in range(self.num_epochs):
            print(f'Epoch {epoch} / {self.num_epochs - 1}')
            print('-' * 20)

            # Each epoch has a training and validation phase:
            for phase in ['train', 'val']:
                if phase == 'train':
                    self.model.train()  # Set model to training mode
                else:
                    self.model.eval()  # Set model to evaluate mode

                phase_loss = 0.0  # Train or val loss
                phase_metric = 0.0
                phase_recall = 0.0
                phase_precision = 0.0

                # Track history only if in train phase:
                with torch.set_grad_enabled(phase == 'train'):
                    # Iterate over data batches:
                    batch = 0
                    for sample in self.dataloaders[phase]:
                        seg_header = sample['header']

                        input1, target1 = sample['input'], sample['target']
                        input, target = input1.to(self.device), target1.to(self.device)
                        # ss = target.shape
                        # ss = np.array(ss)
                        # print(ss)

                        # 1x2x144x144x144

                        # Forward pass:
                        output = self.model(input)
                        # output_ = output_.detach().cpu().numpy()
                        # output = skTrans.resize(output_, ss, order=2, preserve_range=True)
                        #
                        # output = torch.from_numpy(output).to(self.device)

                        loss = self.criterion(output, target)
                        metric = self.metric(output,target)
                        recall = self.recall(output,target)
                        precision = self.precision(output,target)
                        # di, fp, fn = self.metric2(output,target,seg_header)



                        # Losses and metric:
                        phase_loss += loss.item()
                        phase_metric += metric.item()
                        phase_recall += recall.item()
                        phase_precision += precision.item()

                        with np.printoptions(precision=3, suppress=True):
                            print(f'batch: {batch} \tbatch loss: \t{loss:.3f} \tmetric: \t{metric:.3f} \trecall: \t{recall:.3f} \tprecision: \t{precision:.3f}')



                        del input, target, output, metric

                        # Backward pass + optimize only if in training phase:
                        if phase == 'train':
                            loss.backward()
                            self.optimizer.step()

                            # zero the parameter gradients:
                            self.optimizer.zero_grad()

                            if self.scheduler and not self.scheduler_step_per_epoch:
                                self.scheduler.step()

                        del loss
                        batch += 1

                phase_loss /= len(self.dataloaders[phase])
                phase_metric /= len(self.dataloaders[phase])
                phase_recall /= len(self.dataloaders[phase])
                phase_precision /= len(self.dataloaders[phase])
                self.learning_curves['loss'][phase].append(phase_loss)
                self.learning_curves['metric'][phase].append(phase_metric)
                self.learning_curves['recall'][phase].append(phase_recall)
                self.learning_curves['precision'][phase].append(phase_precision)

                print(f'{phase.upper()} \tloss: \t{phase_loss:.3f} \tavg_metric: \t{np.mean(phase_metric):.3f} \trecall: \t{phase_recall:.3f} \tprecision: \t{phase_precision:.3f}')

                wb = openpyxl.load_workbook(self.filepath)
                ws = wb.active
                if phase == 'train':
                    ws[str('A') + str(epoch + 1)] = epoch
                    ws[str('B') + str(epoch + 1)] = phase_loss
                    ws[str('C') + str(epoch + 1)] = phase_metric
                    ws[str('H') + str(epoch + 1)] = phase_recall
                    ws[str('I') + str(epoch + 1)] = phase_precision

                else:
                    ws[str('E') + str(epoch + 1)] = phase_loss
                    ws[str('F') + str(epoch + 1)] = phase_metric
                    ws[str('K') + str(epoch + 1)] = phase_recall
                    ws[str('L') + str(epoch + 1)] = phase_precision

                wb.save(self.filepath)

                # Save summary if it is the best val results so far:
                if phase == 'val':
                    if self.mode == 'max' and np.mean(phase_metric) > self.best_val_avg_metric:
                        self.best_val_epoch = epoch
                        self.best_val_avg_metric = np.mean(phase_metric)
                        self.best_val_loss = phase_loss
                        self.best_val_avg_metric = np.mean(phase_metric)
                        self.best_val_metric = phase_metric
                        self.best_model_wts = copy.deepcopy(self.model.state_dict())

                    if self.mode == 'min' and np.mean(phase_metric) < self.best_val_avg_metric:
                        self.best_val_epoch = epoch
                        self.best_val_avg_metric = np.mean(phase_metric)
                        self.best_val_loss = phase_loss
                        self.best_val_avg_metric = np.mean(phase_metric)
                        self.best_val_metric = phase_metric
                        self.best_model_wts = copy.deepcopy(self.model.state_dict())



            # Adjust learning rate after val phase:
            if self.scheduler and self.scheduler_step_per_epoch:
                if isinstance(self.scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau):
                    self.scheduler.step(np.mean(phase_metric))
                else:
                    self.scheduler.step()

            if (epoch+1) % 10 == 0:
                torch.save(self.best_model_wts, os.path.join(path_to_dir, str(epoch)+'_'+str(self.best_val_epoch)+'_best_model_weights.pt'))

            if (epoch+1) % 10 == 0:
                torch.save(self.model.state_dict(), os.path.join(path_to_dir, str(epoch) +'_weights.pt'))

        if self.save_last_model:
            self.checkpoint = {'model_state_dict': copy.deepcopy(self.model.state_dict()),
                               'optimizer_state_dict': copy.deepcopy(self.optimizer.state_dict())}

    def save_results(self, path_to_dir):
        """"
        Save results in a directory. The method must be used after training.

        A short summary is stored in a csv file ('summary.csv'). Weights of the best model are stored in
        'best_model_weights.pt'. A checkpoint of the last epoch is stored in 'last_model_checkpoint.tar'. Two plots
        for the loss function and metric are stored in 'loss_plot.png' and 'metric_plot.png', respectively.

        Parameters
        ----------
        path_to_dir : str
            A path to the directory for storing all results.
        """

        path_to_dir = pathlib.Path(path_to_dir)

        # Check if the directory exists:
        if not os.path.exists(path_to_dir):
            os.makedirs(path_to_dir)

        # Write a short summary in a csv file:
        with open(path_to_dir / 'summary.csv', 'w', newline='', encoding='utf-8') as summary:
            summary.write(f'SUMMARY OF THE EXPERIMENT:\n\n')
            summary.write(f'BEST VAL EPOCH: {self.best_val_epoch}\n')
            summary.write(f'BEST VAL AVG metric: {self.best_val_avg_metric}\n')
            summary.write(f'BEST VAL LOSS: {self.best_val_loss}\n')
            summary.write(f'BEST VAL AVG metric: {self.best_val_avg_metric}\n')
            summary.write(f'BEST VAL metric: {self.best_val_metric}\n')

        # Save best model weights:
        torch.save(self.best_model_wts, path_to_dir / 'best_model_weights.pt')

        # Save last model weights (checkpoint):
        if self.save_last_model:
            torch.save(self.checkpoint, path_to_dir / 'last_model_checkpoint.tar')

        # Save learning curves as pandas df:
        df_learning_curves = pd.DataFrame.from_dict({
            'loss_train': self.learning_curves['loss']['train'],
            'loss_val': self.learning_curves['loss']['val'],
            'metric_train': self.learning_curves['metric']['train'],
            'metric_val': self.learning_curves['metric']['val']
        })
        df_learning_curves.to_csv(path_to_dir / 'learning_curves.csv', sep=';')

        # Save learning curves' plots in png files:
        # Loss figure:
        plt.figure(figsize=(17.5, 10))
        plt.plot(range(self.num_epochs), self.learning_curves['loss']['train'], label='train')
        plt.plot(range(self.num_epochs), self.learning_curves['loss']['val'], label='val')
        plt.xlabel('Epoch', fontsize=20)
        plt.ylabel('Loss', fontsize=20)
        plt.xticks(fontsize=15)
        plt.yticks(fontsize=15)
        plt.legend(fontsize=20)
        plt.grid()
        plt.savefig(path_to_dir / 'loss_plot.png', bbox_inches='tight')

        # metric figure:
        train_avg_metric = [np.mean(i) for i in self.learning_curves['metric']['train']]
        val_avg_metric = [np.mean(i) for i in self.learning_curves['metric']['val']]

        plt.figure(figsize=(17.5, 10))
        plt.plot(range(self.num_epochs), train_avg_metric, label='train')
        plt.plot(range(self.num_epochs), val_avg_metric, label='val')
        plt.xlabel('Epoch', fontsize=20)
        plt.ylabel('Avg metric', fontsize=20)
        plt.xticks(fontsize=15)
        plt.yticks(fontsize=15)
        plt.legend(fontsize=20)
        plt.grid()
        plt.savefig(path_to_dir / 'metric_plot.png', bbox_inches='tight')

        # metric figure:
        train_avg_recall = [np.mean(i) for i in self.learning_curves['recall']['train']]
        val_avg_recall = [np.mean(i) for i in self.learning_curves['recall']['val']]

        plt.figure(figsize=(17.5, 10))
        plt.plot(range(self.num_epochs), train_avg_recall, label='train')
        plt.plot(range(self.num_epochs), val_avg_recall, label='val')
        plt.xlabel('Epoch', fontsize=20)
        plt.ylabel('Avg recall', fontsize=20)
        plt.xticks(fontsize=15)
        plt.yticks(fontsize=15)
        plt.legend(fontsize=20)
        plt.grid()
        plt.savefig(path_to_dir / 'recall_plot.png', bbox_inches='tight')

        # metric figure:
        train_avg_precision = [np.mean(i) for i in self.learning_curves['precision']['train']]
        val_avg_precision = [np.mean(i) for i in self.learning_curves['precision']['val']]

        plt.figure(figsize=(17.5, 10))
        plt.plot(range(self.num_epochs), train_avg_precision, label='train')
        plt.plot(range(self.num_epochs), val_avg_precision, label='val')
        plt.xlabel('Epoch', fontsize=20)
        plt.ylabel('Avg precision', fontsize=20)
        plt.xticks(fontsize=15)
        plt.yticks(fontsize=15)
        plt.legend(fontsize=20)
        plt.grid()
        plt.savefig(path_to_dir / 'precision_plot.png', bbox_inches='tight')

        print(f'All results have been saved in {path_to_dir}')
